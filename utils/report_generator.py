"""
utils/report_generator.py — SecureGuard AI v2.0
=================================================
Enterprise cybersecurity report generator.

Exports
-------
generate_pdf_report(user_name, user_email, analyses, stats) → bytes
generate_csv_export(analyses) → str
generate_excel_export(user_name, analyses, stats) → bytes
"""

from __future__ import annotations

import csv
import io
import logging
import os
import tempfile
import uuid
from collections import Counter
from datetime import datetime
from typing import Any

# ── ReportLab ──────────────────────────────────────────────────────────────
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    HRFlowable, Image, PageBreak, Paragraph,
    SimpleDocTemplate, Spacer, Table, TableStyle,
)

# ── Matplotlib (charts embedded in PDF) ───────────────────────────────────
import matplotlib
matplotlib.use("Agg")          # non-interactive backend, no display needed
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

# ── openpyxl (Excel) ──────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# ── Brand colours ─────────────────────────────────────────────────────────
C_BG      = colors.HexColor("#0f172a")
C_ACCENT  = colors.HexColor("#00e5ff")
C_PURPLE  = colors.HexColor("#7c4dff")
C_SUCCESS = colors.HexColor("#2ecc71")
C_WARNING = colors.HexColor("#f59e0b")
C_DANGER  = colors.HexColor("#ef4444")
C_MUTED   = colors.HexColor("#94a3b8")
C_TEXT    = colors.HexColor("#f8fafc")
C_SUBTEXT = colors.HexColor("#cbd5e1")
C_PANEL   = colors.HexColor("#1e293b")
C_BORDER  = colors.HexColor("#334155")

STRENGTH_COL: dict[str, Any] = {
    "Very Weak": C_DANGER,  "Weak": C_WARNING,
    "Moderate":  colors.HexColor("#3b82f6"),
    "Strong":    C_PURPLE,  "Excellent": C_SUCCESS,
}
SOURCE_COL: dict[str, Any] = {
    "Hybrid Consensus": C_SUCCESS,  "Machine Learning": C_PURPLE,
    "Rule Engine":      C_WARNING,  "Weighted Decision": colors.HexColor("#3b82f6"),
    "Breach Guard":     C_DANGER,
}

# Matplotlib hex strings (separate from ReportLab colours)
MP_STRENGTH = {
    "Very Weak": "#ef4444", "Weak": "#f59e0b",
    "Moderate": "#3b82f6",  "Strong": "#7c4dff", "Excellent": "#2ecc71",
}
MP_SOURCE = {
    "Hybrid Consensus": "#2ecc71", "Machine Learning": "#7c4dff",
    "Rule Engine": "#f59e0b",      "Weighted Decision": "#3b82f6",
    "Breach Guard": "#ef4444",
}
MP_AGREE = {"Agreed": "#2ecc71", "Disagreed": "#ef4444"}
APP_VERSION = "2.0"


# ── Style factory ─────────────────────────────────────────────────────────

def _styles() -> dict[str, ParagraphStyle]:
    def _s(n, **kw): return ParagraphStyle(n, **kw)
    return {
        "cover_title": _s("CT", fontSize=30, textColor=C_ACCENT,
                           fontName="Helvetica-Bold", alignment=TA_CENTER, spaceAfter=4),
        "cover_sub":   _s("CS", fontSize=13, textColor=C_SUBTEXT,
                           fontName="Helvetica", alignment=TA_CENTER, spaceAfter=3),
        "cover_meta":  _s("CM", fontSize=10, textColor=C_MUTED,
                           fontName="Helvetica", alignment=TA_CENTER, spaceAfter=2),
        "h1":  _s("H1", fontSize=17, textColor=C_ACCENT,
                   fontName="Helvetica-Bold", spaceBefore=16, spaceAfter=7),
        "h2":  _s("H2", fontSize=12, textColor=C_SUBTEXT,
                   fontName="Helvetica-Bold", spaceBefore=9, spaceAfter=4),
        "body": _s("B",  fontSize=9.5, textColor=C_SUBTEXT,
                    fontName="Helvetica", leading=14, spaceAfter=3),
        "small": _s("S", fontSize=8.5, textColor=C_MUTED,
                     fontName="Helvetica", leading=12),
        "bullet": _s("BU", fontSize=9.5, textColor=C_SUBTEXT,
                      fontName="Helvetica", leftIndent=10, leading=14, spaceAfter=2),
        "footer": _s("FT", fontSize=7.5, textColor=C_MUTED,
                      fontName="Helvetica", alignment=TA_CENTER),
        "conf":   _s("CF", fontSize=8.5, textColor=C_DANGER,
                      fontName="Helvetica-Bold", alignment=TA_CENTER),
    }


def _hr(c=C_BORDER, w=0.4):
    return HRFlowable(width="100%", thickness=w, color=c, spaceAfter=5, spaceBefore=2)


def _tbl_base() -> list:
    return [
        ("BACKGROUND",    (0,0), (-1,0),  C_PANEL),
        ("TEXTCOLOR",     (0,0), (-1,0),  C_ACCENT),
        ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0),  8.5),
        ("BOTTOMPADDING", (0,0), (-1,0),  7),
        ("TOPPADDING",    (0,0), (-1,0),  7),
        ("BACKGROUND",    (0,1), (-1,-1), C_BG),
        ("TEXTCOLOR",     (0,1), (-1,-1), C_SUBTEXT),
        ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",      (0,1), (-1,-1), 8.5),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [C_BG, C_PANEL]),
        ("GRID",          (0,0), (-1,-1), 0.25, C_BORDER),
        ("ROWPADDING",    (0,0), (-1,-1), 5),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
    ]


# ── Page canvas callbacks ─────────────────────────────────────────────────

def _bg(canvas, doc):
    w, h = A4
    canvas.saveState()
    canvas.setFillColor(C_BG)
    canvas.rect(0, 0, w, h, fill=1, stroke=0)
    canvas.setFillColor(C_ACCENT)
    canvas.rect(0, 0, 3.5, h, fill=1, stroke=0)
    canvas.restoreState()


def _bg_footer(canvas, doc):
    _bg(canvas, doc)
    w, h = A4
    canvas.saveState()
    canvas.setFont("Helvetica", 7.5)
    canvas.setFillColor(C_MUTED)
    canvas.drawCentredString(
        w / 2, 11 * mm,
        f"SecureGuard AI v{APP_VERSION}  —  CONFIDENTIAL  |  Page {doc.page}",
    )
    canvas.restoreState()


# ── Chart helpers (matplotlib → temp PNG → ReportLab Image) ───────────────

def _save_tmp(fig) -> str:
    """Save a matplotlib figure to a temp file; return the path."""
    fd, path = tempfile.mkstemp(suffix=".png")
    os.close(fd)
    fig.savefig(path, dpi=130, bbox_inches="tight",
                facecolor="#0f172a", edgecolor="none")
    plt.close(fig)
    return path


def _chart_strength_dist(analyses: list[dict]) -> str:
    """Pie chart: strength category distribution."""
    counts = Counter(r.get("strength_category", "Unknown") for r in analyses)
    order  = ["Very Weak", "Weak", "Moderate", "Strong", "Excellent"]
    labels = [k for k in order if counts.get(k, 0) > 0]
    sizes  = [counts[k] for k in labels]
    clrs   = [MP_STRENGTH.get(k, "#94a3b8") for k in labels]

    fig, ax = plt.subplots(figsize=(4.5, 3.5), facecolor="#0f172a")
    wedges, texts, autotexts = ax.pie(
        sizes, labels=None, colors=clrs, autopct="%1.0f%%",
        startangle=90, pctdistance=0.75,
        wedgeprops=dict(width=0.6, linewidth=0.5, edgecolor="#0f172a"),
    )
    for at in autotexts:
        at.set(color="white", fontsize=8, fontweight="bold")
    ax.legend(
        wedges, labels,
        loc="lower center", ncol=3, frameon=False,
        labelcolor="white", fontsize=8,
        bbox_to_anchor=(0.5, -0.18),
    )
    ax.set_title("Password Strength Distribution",
                 color="#00e5ff", fontsize=10, pad=8)
    return _save_tmp(fig)


def _chart_hybrid_dist(analyses: list[dict]) -> str:
    """Bar chart: hybrid decision distribution."""
    order  = ["Very Weak", "Weak", "Moderate", "Strong", "Excellent"]
    counts = Counter(r.get("hybrid_decision") for r in analyses
                     if r.get("hybrid_decision"))
    vals   = [counts.get(k, 0) for k in order]
    clrs   = [MP_STRENGTH.get(k, "#94a3b8") for k in order]

    fig, ax = plt.subplots(figsize=(5, 3.5), facecolor="#0f172a")
    ax.set_facecolor("#1e293b")
    bars = ax.bar(order, vals, color=clrs, width=0.5,
                  linewidth=0, zorder=3)
    ax.bar_label(bars, padding=3, color="white", fontsize=8, fontweight="bold")
    ax.set_ylabel("Count", color="#cbd5e1", fontsize=8)
    ax.tick_params(colors="#cbd5e1", labelsize=8)
    ax.spines[:].set_visible(False)
    ax.yaxis.grid(True, color="#334155", linewidth=0.4, zorder=0)
    ax.set_title("Hybrid Decision Distribution",
                 color="#00e5ff", fontsize=10, pad=8)
    return _save_tmp(fig)


def _chart_agreement(analyses: list[dict]) -> str:
    """Donut chart: rule vs ML agreement."""
    agreed    = sum(1 for r in analyses if r.get("hybrid_agreement") is True)
    disagreed = sum(1 for r in analyses if r.get("hybrid_agreement") is False)
    if agreed + disagreed == 0:
        agreed = 1   # avoid empty chart

    fig, ax = plt.subplots(figsize=(4, 3.5), facecolor="#0f172a")
    wedges, _, auto = ax.pie(
        [agreed, disagreed],
        labels=None,
        colors=["#2ecc71", "#ef4444"],
        autopct="%1.0f%%",
        startangle=90,
        pctdistance=0.72,
        wedgeprops=dict(width=0.55, linewidth=0.5, edgecolor="#0f172a"),
    )
    for at in auto:
        at.set(color="white", fontsize=9, fontweight="bold")
    patches = [
        mpatches.Patch(color="#2ecc71", label=f"Agreed ({agreed})"),
        mpatches.Patch(color="#ef4444", label=f"Disagreed ({disagreed})"),
    ]
    ax.legend(handles=patches, loc="lower center", frameon=False,
              labelcolor="white", fontsize=8, bbox_to_anchor=(0.5, -0.15))
    ax.set_title("Rule vs ML Agreement", color="#00e5ff", fontsize=10, pad=8)
    return _save_tmp(fig)


def _chart_avg_confidence(analyses: list[dict]) -> str:
    """Horizontal bar chart: average ML confidence per hybrid decision."""
    order = ["Very Weak", "Weak", "Moderate", "Strong", "Excellent"]
    sums  = {k: [] for k in order}
    for r in analyses:
        dec  = r.get("hybrid_decision")
        conf = r.get("ml_confidence")
        if dec in sums and conf is not None:
            sums[dec].append(float(conf))
    avgs  = [sum(sums[k]) / len(sums[k]) * 100 if sums[k] else 0 for k in order]
    clrs  = [MP_STRENGTH.get(k, "#94a3b8") for k in order]

    fig, ax = plt.subplots(figsize=(5, 3.5), facecolor="#0f172a")
    ax.set_facecolor("#1e293b")
    bars = ax.barh(order, avgs, color=clrs, height=0.4, linewidth=0, zorder=3)
    ax.bar_label(bars, fmt="%.0f%%", padding=4, color="white",
                 fontsize=8, fontweight="bold")
    ax.set_xlim(0, 115)
    ax.tick_params(colors="#cbd5e1", labelsize=8)
    ax.spines[:].set_visible(False)
    ax.xaxis.grid(True, color="#334155", linewidth=0.4, zorder=0)
    ax.set_title("Avg ML Confidence per Decision",
                 color="#00e5ff", fontsize=10, pad=8)
    return _save_tmp(fig)


# ── PDF section builders ──────────────────────────────────────────────────

def _page1_cover(s, user_name, user_email, report_id, ts) -> list:
    out = [Spacer(1, 45*mm)]
    out.append(Paragraph("SecureGuard AI", s["cover_title"]))
    out.append(Paragraph("Cybersecurity Password Assessment Report", s["cover_sub"]))
    out.append(Spacer(1, 8*mm))
    for txt in [
        f"Prepared for:  {user_name}",
        f"Email:         {user_email}",
        f"Generated:     {ts}",
        f"Report ID:     {report_id}",
        f"Version:       {APP_VERSION}",
    ]:
        out.append(Paragraph(txt, s["cover_meta"]))
    out.append(Spacer(1, 8*mm))
    out.append(Paragraph(
        "CONFIDENTIAL — For authorised personnel only.", s["conf"],
    ))
    out.append(PageBreak())
    return out


def _page2_exec_summary(s, stats, analyses) -> list:
    out = [Paragraph("1. Executive Summary", s["h1"]), _hr()]

    avg_score = stats.get("avg_score", 0.0)
    total     = stats.get("total", 0)
    breached  = stats.get("breached", 0)
    threat    = stats.get("threat_level", "N/A")

    # Overall rating
    if avg_score >= 75:
        rating, rc = "GOOD",     C_SUCCESS
    elif avg_score >= 55:
        rating, rc = "FAIR",     colors.HexColor("#3b82f6")
    elif avg_score >= 35:
        rating, rc = "POOR",     C_WARNING
    else:
        rating, rc = "CRITICAL", C_DANGER

    out.append(Paragraph(
        f'Overall Security Rating: <font color="#{rc.hexval()[2:]}">'
        f'<b>{rating}</b></font>',
        ParagraphStyle("RatingP", fontSize=13, textColor=C_TEXT,
                       fontName="Helvetica-Bold", spaceAfter=7),
    ))

    # Get most recent analysis for hybrid fields
    primary = analyses[0] if analyses else {}
    hybrid_dec  = primary.get("hybrid_decision") or "N/A"
    ml_conf     = primary.get("ml_confidence")
    conf_str    = f"{ml_conf*100:.1f}%" if ml_conf is not None else "N/A"
    breach_stat = "COMPROMISED" if primary.get("is_breached") else "Safe"

    data = [
        ["KPI", "Value", "Assessment"],
        ["Overall Score",       f"{avg_score}/100",  rating],
        ["Threat Level",        threat,               threat],
        ["Hybrid AI Decision",  hybrid_dec,           hybrid_dec],
        ["ML Confidence",       conf_str,             "—"],
        ["Breach Status",       breach_stat,          "CRITICAL" if breached > 0 else "Safe"],
        ["Total Analyses",      str(total),           "—"],
    ]
    tbl = Table(data, colWidths=[60*mm, 55*mm, 55*mm])
    ts  = _tbl_base()
    status_map = {
        "GOOD":C_SUCCESS,"FAIR":colors.HexColor("#3b82f6"),
        "POOR":C_WARNING,"CRITICAL":C_DANGER,
        "Safe":C_SUCCESS,"COMPROMISED":C_DANGER,
    }
    for i, row in enumerate(data[1:], 1):
        for col_idx in (2,):
            col = status_map.get(row[col_idx], C_MUTED)
            ts.append(("TEXTCOLOR", (col_idx, i), (col_idx, i), col))
            ts.append(("FONTNAME",  (col_idx, i), (col_idx, i), "Helvetica-Bold"))
        # Colour hybrid decision
        c2 = STRENGTH_COL.get(row[1], None)
        if c2 and row[0] in ("Hybrid AI Decision", "Threat Level"):
            ts.append(("TEXTCOLOR", (1, i), (1, i), c2))

    tbl.setStyle(TableStyle(ts))
    out.append(tbl)
    out.append(Spacer(1, 6*mm))
    return out


def _page3_password_analysis(s, analyses) -> list:
    out = [PageBreak(), Paragraph("2. Password Analysis", s["h1"]), _hr()]
    if not analyses:
        out.append(Paragraph("No analysis records available.", s["body"]))
        return out

    out.append(Paragraph(
        f"Showing {min(len(analyses), 30)} of {len(analyses)} records (newest first).",
        s["small"],
    ))
    out.append(Spacer(1, 3*mm))

    headers = ["#", "Label", "Category", "Score", "Entropy",
               "Len", "Up", "Lo", "Dig", "Sp",
               "Dict", "Kbd", "Breached"]
    rows_data = [headers]
    for i, r in enumerate(analyses[:30], 1):
        agr_flag = "Y" if r.get("is_breached") else "N"
        rows_data.append([
            str(i),
            r.get("label", "—"),
            r.get("strength_category", "—"),
            str(r.get("strength_score", "—")),
            str(r.get("entropy", "—")),
            str(r.get("length", "—")),
            str(r.get("uppercase", 0)),
            str(r.get("lowercase", 0)),
            str(r.get("digits", 0)),
            str(r.get("special", 0)),
            "Y" if r.get("dictionary_word") else "N",
            "Y" if r.get("keyboard_pattern") else "N",
            agr_flag,
        ])

    widths = [8, 28, 22, 15, 16, 10, 9, 9, 9, 9, 11, 11, 16]
    widths = [w * mm for w in widths]
    tbl = Table(rows_data, colWidths=widths, repeatRows=1)
    ts  = _tbl_base()
    ts.append(("FONTSIZE", (0,0), (-1,-1), 7.5))
    # Colour category column
    for i, r in enumerate(analyses[:30], 1):
        cat_c = STRENGTH_COL.get(r.get("strength_category",""), C_MUTED)
        ts.append(("TEXTCOLOR", (2, i), (2, i), cat_c))
        ts.append(("FONTNAME",  (2, i), (2, i), "Helvetica-Bold"))
        if r.get("is_breached"):
            ts.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#2d1515")))
            ts.append(("TEXTCOLOR",  (12, i),(12, i), C_DANGER))
        if r.get("dictionary_word"):
            ts.append(("TEXTCOLOR", (10, i),(10, i), C_DANGER))
        if r.get("keyboard_pattern"):
            ts.append(("TEXTCOLOR", (11, i),(11, i), C_DANGER))

    tbl.setStyle(TableStyle(ts))
    out.append(tbl)
    out.append(Spacer(1, 5*mm))
    return out


def _page4_hybrid_ai(s, analyses) -> list:
    out = [PageBreak(), Paragraph("3. Hybrid AI Analysis", s["h1"]), _hr()]
    hybrid_rows = [r for r in analyses if r.get("hybrid_decision")]
    if not hybrid_rows:
        out.append(Paragraph("No Hybrid AI analyses recorded yet.", s["body"]))
        return out

    out.append(Paragraph(
        f"{len(hybrid_rows)} records with Hybrid AI data (newest first, max 25).",
        s["small"],
    ))
    out.append(Spacer(1, 3*mm))

    headers = ["#", "Label", "Rule", "ML Pred", "Conf%",
               "Hybrid", "Agree", "Source"]
    rows_data = [headers]
    for i, r in enumerate(hybrid_rows[:25], 1):
        conf = r.get("ml_confidence")
        conf_str = f"{conf*100:.0f}" if conf is not None else "—"
        agr  = r.get("hybrid_agreement")
        agr_str = "Yes" if agr is True else ("No" if agr is False else "—")
        rows_data.append([
            str(i),
            r.get("label", "—"),
            r.get("strength_category", "—"),
            r.get("ml_prediction") or "—",
            conf_str,
            r.get("hybrid_decision") or "—",
            agr_str,
            r.get("decision_source") or "—",
        ])

    widths = [8, 28, 22, 22, 14, 22, 14, 38]
    widths = [w * mm for w in widths]
    tbl = Table(rows_data, colWidths=widths, repeatRows=1)
    ts  = _tbl_base()
    ts.append(("FONTSIZE", (0,0), (-1,-1), 7.5))
    for i, r in enumerate(hybrid_rows[:25], 1):
        # Rule colour
        ts.append(("TEXTCOLOR", (2,i),(2,i), STRENGTH_COL.get(r.get("strength_category",""), C_MUTED)))
        ts.append(("FONTNAME",  (2,i),(2,i), "Helvetica-Bold"))
        # ML colour
        ts.append(("TEXTCOLOR", (3,i),(3,i), STRENGTH_COL.get(r.get("ml_prediction",""), C_MUTED)))
        # Hybrid colour
        ts.append(("TEXTCOLOR", (5,i),(5,i), STRENGTH_COL.get(r.get("hybrid_decision",""), C_MUTED)))
        ts.append(("FONTNAME",  (5,i),(5,i), "Helvetica-Bold"))
        # Agreement colour
        agr = r.get("hybrid_agreement")
        if agr is True:
            ts.append(("TEXTCOLOR", (6,i),(6,i), C_SUCCESS))
        elif agr is False:
            ts.append(("TEXTCOLOR", (6,i),(6,i), C_WARNING))
        # Source colour
        ts.append(("TEXTCOLOR", (7,i),(7,i), SOURCE_COL.get(r.get("decision_source",""), C_MUTED)))

    tbl.setStyle(TableStyle(ts))
    out.append(tbl)

    # Show reason for the most-recent record
    if hybrid_rows[0].get("hybrid_reason"):
        out.append(Spacer(1, 4*mm))
        out.append(Paragraph("Most Recent AI Reason:", s["h2"]))
        out.append(Paragraph(hybrid_rows[0]["hybrid_reason"], s["body"]))

    out.append(Spacer(1, 5*mm))
    return out


def _page5_breach_intelligence(s, analyses) -> list:
    out = [PageBreak(), Paragraph("4. Breach Intelligence", s["h1"]), _hr()]
    breached_rows = [r for r in analyses if r.get("is_breached")]
    total = len(analyses)

    breach_count = len(breached_rows)
    breach_rate  = round(breach_count / total * 100, 1) if total else 0.0
    status_str   = f"{breach_count} breached ({breach_rate}%)" if breach_count else "No breaches detected"
    status_col   = C_DANGER if breach_count > 0 else C_SUCCESS

    out.append(Paragraph(
        f'Overall Breach Status: <font color="#{status_col.hexval()[2:]}">'
        f'<b>{status_str}</b></font>',
        s["body"],
    ))
    out.append(Spacer(1, 3*mm))

    summary_data = [
        ["Metric", "Value"],
        ["Total Analyses",   str(total)],
        ["Breached Found",   str(breach_count)],
        ["Breach Rate",      f"{breach_rate}%"],
        ["Breach Database",  "RockYou Sample (437 common passwords)"],
        ["Overall Status",   "COMPROMISED" if breach_count > 0 else "Safe"],
    ]
    tbl = Table(summary_data, colWidths=[80*mm, 85*mm])
    ts  = _tbl_base()
    if breach_count > 0:
        ts.append(("TEXTCOLOR", (1, 5), (1, 5), C_DANGER))
        ts.append(("FONTNAME",  (1, 5), (1, 5), "Helvetica-Bold"))
    tbl.setStyle(TableStyle(ts))
    out.append(tbl)

    out.append(Spacer(1, 4*mm))
    out.append(Paragraph("Breach Details:", s["h2"]))

    if breached_rows:
        b_headers = ["#", "Label", "Strength", "Score", "Entropy", "Risk"]
        b_data    = [b_headers]
        for i, r in enumerate(breached_rows[:15], 1):
            b_data.append([
                str(i),
                r.get("label", "—"),
                r.get("strength_category", "—"),
                str(r.get("strength_score", "—")),
                str(r.get("entropy", "—")),
                "Critical",
            ])
        tbl2 = Table(b_data, colWidths=[10,35,28,16,16,20])
        tbl2 = Table(b_data, colWidths=[10*mm,35*mm,28*mm,16*mm,16*mm,20*mm])
        ts2  = _tbl_base()
        ts2.append(("FONTSIZE", (0,0),(-1,-1), 8))
        for i in range(1, len(b_data)):
            ts2.append(("BACKGROUND",(0,i),(-1,i), colors.HexColor("#2d1515")))
            ts2.append(("TEXTCOLOR", (5,i),(5,i),  C_DANGER))
            ts2.append(("FONTNAME",  (5,i),(5,i),  "Helvetica-Bold"))
        tbl2.setStyle(TableStyle(ts2))
        out.append(tbl2)
    else:
        out.append(Paragraph(
            "No breached passwords detected. All analysed passwords are breach-free.",
            s["body"],
        ))

    out.append(Spacer(1, 4*mm))
    out.append(Paragraph("Security Impact & Recommended Actions:", s["h2"]))
    if breach_count > 0:
        actions = [
            "CRITICAL: Change breached passwords immediately on all affected services.",
            "Enable Multi-Factor Authentication on accounts using these passwords.",
            "Never reuse a compromised password on any service.",
            "Conduct a full credential audit across all organisational systems.",
            "Implement automated breach monitoring for all new passwords.",
        ]
    else:
        actions = [
            "No breaches detected. Continue regular scanning for ongoing protection.",
            "Enable MFA on all accounts as a proactive security measure.",
            "Schedule periodic breach scans for all critical service passwords.",
            "Monitor HaveIBeenPwned for newly disclosed breach datasets.",
        ]
    for a in actions:
        out.append(Paragraph(f"• {a}", s["bullet"]))

    out.append(Spacer(1, 5*mm))
    return out

def _page6_charts(s, analyses, tmp_files: list) -> list:
    """Page 5 — four embedded matplotlib charts."""
    out = [PageBreak(), Paragraph("5. Visual Analytics", s["h1"]), _hr()]

    if not analyses:
        out.append(Paragraph("No data available for chart generation.", s["body"]))
        return out

    chart_funcs = [
        ("Password Strength Distribution",  _chart_strength_dist),
        ("Hybrid Decision Distribution",     _chart_hybrid_dist),
        ("Rule vs ML Agreement",             _chart_agreement),
        ("Average ML Confidence per Tier",   _chart_avg_confidence),
    ]

    # Render charts in a 2x2 grid using a Table
    chart_cells = []
    row_pair = []
    for title, fn in chart_funcs:
        try:
            path = fn(analyses)
            tmp_files.append(path)
            img = Image(path, width=84*mm, height=62*mm)
            cell_content = [
                Paragraph(title, ParagraphStyle(
                    "ChartTitle", fontSize=8.5, textColor=C_SUBTEXT,
                    fontName="Helvetica-Bold", alignment=TA_CENTER, spaceAfter=3,
                )),
                img,
            ]
        except Exception as exc:
            log.warning("Chart generation failed for %r: %s", title, exc)
            cell_content = [Paragraph(f"[{title}: unavailable]", s["small"])]

        row_pair.append(cell_content)
        if len(row_pair) == 2:
            chart_cells.append(row_pair)
            row_pair = []

    if row_pair:                          # odd chart count
        row_pair.append([Paragraph("", s["small"])])
        chart_cells.append(row_pair)

    if chart_cells:
        grid = Table(chart_cells, colWidths=[90*mm, 90*mm])
        grid.setStyle(TableStyle([
            ("VALIGN",  (0,0), (-1,-1), "TOP"),
            ("ALIGN",   (0,0), (-1,-1), "CENTER"),
            ("PADDING", (0,0), (-1,-1), 4),
        ]))
        out.append(grid)

    out.append(Spacer(1, 5*mm))
    return out


def _page7_recommendations(s, analyses) -> list:
    """Page 6 — AI Security Recommendations."""
    out = [PageBreak(), Paragraph("6. AI Security Recommendations", s["h1"]), _hr()]

    primary   = analyses[0] if analyses else {}
    score     = primary.get("strength_score", 0)
    is_breach = primary.get("is_breached", False)
    has_dict  = primary.get("dictionary_word", False)
    has_kbd   = primary.get("keyboard_pattern", False)
    length    = primary.get("length", 0)
    special   = primary.get("special", 0)
    uppercase = primary.get("uppercase", 0)
    digits    = primary.get("digits", 0)

    # ── Immediate actions ──────────────────────────────────────────────
    out.append(Paragraph("Immediate Actions", s["h2"]))
    immediate = []
    if is_breach:
        immediate.append("CRITICAL: Change this password on every service immediately.")
    if score < 35:
        immediate.append("Password is critically weak. Replace it before your next login.")
    if has_kbd:
        immediate.append("Remove keyboard sequences (qwerty, 12345) from all passwords.")
    if has_dict:
        immediate.append("Replace dictionary words with random character combinations.")
    if length < 8:
        immediate.append("Increase password length to at least 14 characters.")
    if special == 0:
        immediate.append("Add special characters (!, @, #) to increase complexity.")
    if uppercase == 0:
        immediate.append("Include at least two uppercase letters.")
    if digits == 0:
        immediate.append("Include at least two numeric digits.")
    if not immediate:
        immediate.append("No critical immediate actions required. Maintain current practices.")
    for a in immediate:
        out.append(Paragraph(f"• {a}", s["bullet"]))

    out.append(Spacer(1, 4*mm))

    # ── Password hygiene ───────────────────────────────────────────────
    out.append(Paragraph("Password Hygiene", s["h2"]))
    for tip in [
        "Use a minimum of 14 characters for all passwords.",
        "Combine uppercase, lowercase, digits, and symbols in every password.",
        "Never use the same password across different services.",
        "Avoid personal information: names, birthdays, phone numbers.",
        "Change passwords that are over 90 days old for high-value accounts.",
    ]:
        out.append(Paragraph(f"• {tip}", s["bullet"]))

    out.append(Spacer(1, 4*mm))

    # ── Credential protection ──────────────────────────────────────────
    out.append(Paragraph("Credential Protection", s["h2"]))
    for tip in [
        "Use a reputable password manager: Bitwarden, 1Password, or KeePass.",
        "Enable breach monitoring alerts for all critical accounts.",
        "Store passwords only in encrypted vaults — never in plain text.",
        "Review and revoke unused app permissions and OAuth tokens regularly.",
    ]:
        out.append(Paragraph(f"• {tip}", s["bullet"]))

    out.append(Spacer(1, 4*mm))

    # ── MFA ────────────────────────────────────────────────────────────
    out.append(Paragraph("Multi-Factor Authentication (MFA)", s["h2"]))
    for tip in [
        "Enable MFA on email accounts as the highest priority.",
        "Enable MFA on banking, financial, and payroll services.",
        "Use TOTP apps (Authy, Google Authenticator) over SMS-based OTP.",
        "For enterprise accounts, deploy hardware security keys (FIDO2/YubiKey).",
    ]:
        out.append(Paragraph(f"• {tip}", s["bullet"]))

    out.append(Spacer(1, 4*mm))

    # ── Long-term improvements ─────────────────────────────────────────
    out.append(Paragraph("Long-Term Improvements", s["h2"]))
    for tip in [
        "Educate team members on phishing and social engineering attacks.",
        "Run regular breach scans through SecureGuard AI.",
        "Implement Single Sign-On (SSO) with strong MFA for enterprise systems.",
        "Schedule quarterly password audits and enforce expiry policies.",
        "Monitor the dark web for leaked credentials using automated tooling.",
    ]:
        out.append(Paragraph(f"• {tip}", s["bullet"]))

    out.append(Spacer(1, 5*mm))
    return out


def _page8_best_practices(s) -> list:
    """Page 7 — Security Best Practices Checklist."""
    out = [PageBreak(), Paragraph("7. Security Best Practices Checklist", s["h1"]), _hr()]

    checklist = [
        "Password is 14 or more characters long",
        "Contains uppercase letters (A–Z)",
        "Contains lowercase letters (a–z)",
        "Contains numeric digits (0–9)",
        "Contains special characters (!@#$%^&*)",
        "Not found in any known breach database",
        "No dictionary words or common phrases",
        "No keyboard sequences (qwerty, 12345, asdf)",
        "No personal information (dates, names, phone numbers)",
        "Unique — not reused on any other service",
        "Multi-Factor Authentication (MFA) is enabled",
        "Password is stored in an encrypted password manager",
        "Password has been rotated within the last 90 days",
        "Breach monitoring / alerts are active for this account",
        "Team has completed security awareness training",
        "OAuth tokens and app permissions reviewed in last 30 days",
        "No plaintext passwords stored in code repositories or documents",
    ]

    data = [["#", "Security Control", "Status"]]
    for i, item in enumerate(checklist, 1):
        data.append([str(i), item, "☐  Pending"])

    tbl = Table(data, colWidths=[10*mm, 130*mm, 28*mm])
    ts  = _tbl_base()
    ts.append(("FONTSIZE",     (0,0), (-1,-1), 9))
    ts.append(("TEXTCOLOR",    (2,1), (2,-1),  C_ACCENT))
    ts.append(("ALIGN",        (2,0), (2,-1),  "CENTER"))
    tbl.setStyle(TableStyle(ts))
    out.append(tbl)
    out.append(Spacer(1, 6*mm))
    return out


def _page9_closing(s, report_id, ts) -> list:
    """Page 8 — Closing / footer page."""
    out = [PageBreak(), Spacer(1, 55*mm)]
    out.append(Paragraph("SecureGuard AI", s["cover_title"]))
    out.append(Paragraph("AI-Powered Password Intelligence Platform", s["cover_sub"]))
    out.append(Spacer(1, 8*mm))
    out.append(_hr(C_ACCENT, 1.2))
    out.append(Spacer(1, 4*mm))
    for line in [
        f"Generated: {ts}",
        f"Report ID: {report_id}",
        f"Application Version: {APP_VERSION}",
        "Classification: CONFIDENTIAL",
    ]:
        out.append(Paragraph(line, s["cover_meta"]))
    out.append(Spacer(1, 8*mm))
    out.append(Paragraph(
        "CONFIDENTIAL NOTICE — This report contains sensitive cybersecurity assessment "
        "information. Distribution is restricted to authorised personnel only. "
        "SecureGuard AI accepts no liability for actions taken based solely on this report. "
        "Always complement automated assessments with expert human review.",
        ParagraphStyle("ClosingNotice", fontSize=8, textColor=C_MUTED,
                       fontName="Helvetica", alignment=TA_CENTER, leading=13),
    ))
    return out


# ── Main PDF generator ────────────────────────────────────────────────────

def generate_pdf_report(
    user_name:  str,
    user_email: str,
    analyses:   list[dict[str, Any]],
    stats:      dict[str, Any],
) -> bytes:
    """
    Build a 7-page enterprise cybersecurity PDF report.

    Parameters
    ----------
    user_name   : Recipient's full name
    user_email  : Recipient's email address
    analyses    : list of to_dict() rows, newest-first
    stats       : get_card_stats() result dict

    Returns
    -------
    bytes — raw PDF binary ready to stream via Flask Response
    """
    report_id = str(uuid.uuid4()).upper()[:16]
    ts        = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    s         = _styles()
    tmp_files: list[str] = []

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=18*mm, rightMargin=18*mm,
        topMargin=15*mm,  bottomMargin=18*mm,
        title="SecureGuard AI — Cybersecurity Password Assessment Report",
        author="SecureGuard AI",
    )

    story: list = []
    story += _page1_cover(s, user_name, user_email, report_id, ts)
    story += _page2_exec_summary(s, stats, analyses)
    story += _page3_password_analysis(s, analyses)
    story += _page4_hybrid_ai(s, analyses)
    story += _page5_breach_intelligence(s, analyses)
    story += _page6_charts(s, analyses, tmp_files)
    story += _page7_recommendations(s, analyses)
    story += _page8_best_practices(s)
    story += _page9_closing(s, report_id, ts)

    try:
        doc.build(story, onFirstPage=_bg, onLaterPages=_bg_footer)
    finally:
        # Always clean up temp chart images
        for p in tmp_files:
            try:
                os.unlink(p)
            except OSError:
                pass

    log.info("PDF generated  user=%s  id=%s  pages~9  analyses=%d",
             user_email, report_id, len(analyses))
    return buf.getvalue()


# ── CSV generator ─────────────────────────────────────────────────────────

def generate_csv_export(analyses: list[dict[str, Any]]) -> str:
    """
    Build a comprehensive CSV export of all password analysis records.
    Includes Hybrid AI fields where available (NULL for legacy rows).

    Returns UTF-8 CSV text.
    """
    fieldnames = [
        "id", "analyzed_at", "label",
        "strength_category", "strength_score", "entropy", "length",
        "uppercase", "lowercase", "digits", "special",
        "dictionary_word", "keyboard_pattern", "is_breached",
        "ml_prediction", "ml_confidence",
        "hybrid_decision", "hybrid_agreement",
        "decision_source", "hybrid_reason",
    ]
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames,
                            extrasaction="ignore", lineterminator="\n")
    writer.writeheader()
    for row in analyses:
        flat = dict(row)
        if flat.get("ml_confidence") is not None:
            flat["ml_confidence"] = f"{float(flat['ml_confidence'])*100:.1f}%"
        writer.writerow(flat)

    log.info("CSV export generated  rows=%d", len(analyses))
    return output.getvalue()


# ── Excel generator ───────────────────────────────────────────────────────

def generate_excel_export(
    user_name: str,
    analyses:  list[dict[str, Any]],
    stats:     dict[str, Any],
) -> bytes:
    """
    Build a multi-sheet .xlsx report with professional formatting.

    Sheets
    ------
    1. Executive Summary
    2. Password Analysis
    3. Hybrid AI
    4. Breach Intelligence
    5. Analytics Summary
    6. Recommendations

    Returns bytes — raw .xlsx binary ready to stream via Flask.
    """
    wb = openpyxl.Workbook()

    # ── Shared style helpers ───────────────────────────────────────────
    NAVY  = "0F172A"
    CYAN  = "00E5FF"
    PANEL = "1E293B"
    SUBTEXT = "CBD5E1"
    SUCCESS = "2ECC71"
    DANGER  = "EF4444"
    WARNING = "F59E0B"

    def _hdr_font():   return Font(bold=True, color=CYAN,    size=10)
    def _body_font():  return Font(color=SUBTEXT, size=9)
    def _title_font(): return Font(bold=True, color=CYAN,    size=14)
    def _label_font(): return Font(bold=True, color=SUBTEXT, size=9)

    def _hdr_fill():   return PatternFill("solid", fgColor=PANEL)
    def _body_fill(alt=False):
        return PatternFill("solid", fgColor="17253A" if alt else NAVY)

    def _thin_border():
        side = Side(style="thin", color="334155")
        return Border(left=side, right=side, top=side, bottom=side)

    def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def _left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def _autofit(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    def _write_headers(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font      = _hdr_font()
            cell.fill      = _hdr_fill()
            cell.border    = _thin_border()
            cell.alignment = _center()

    def _style_data_row(ws, row_idx, alt=False, danger=False):
        fill = PatternFill("solid", fgColor="2D1515" if danger else ("17253A" if alt else NAVY))
        for cell in ws[row_idx]:
            cell.fill      = fill
            cell.font      = Font(color=DANGER if danger else SUBTEXT, size=9)
            cell.border    = _thin_border()
            cell.alignment = _left()

    # ── Sheet 1: Executive Summary ─────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.sheet_view.showGridLines = False
    ws1["A1"] = "SecureGuard AI — Cybersecurity Password Assessment Report"
    ws1["A1"].font = _title_font()
    ws1["A2"] = f"Prepared for: {user_name}"
    ws1["A2"].font = _label_font()
    ws1["A3"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    ws1["A3"].font = Font(color=SUBTEXT, size=9)
    ws1.append([])

    kpi_headers = ["KPI", "Value", "Status"]
    ws1.append(kpi_headers)
    for cell in ws1[5]:
        cell.font = _hdr_font(); cell.fill = _hdr_fill()
        cell.border = _thin_border(); cell.alignment = _center()

    primary = analyses[0] if analyses else {}
    ml_conf = primary.get("ml_confidence")
    kpi_rows = [
        ["Total Analyses",     stats.get("total", 0),      "—"],
        ["Average Score",      f"{stats.get('avg_score',0)}/100", "—"],
        ["Strong Passwords",   stats.get("strong", 0),     f"{stats.get('strong_pct',0)}%"],
        ["Weak Passwords",     stats.get("weak", 0),       f"{stats.get('weak_pct',0)}%"],
        ["Breached Found",     stats.get("breached", 0),   "Critical" if stats.get("breached",0) > 0 else "Safe"],
        ["Threat Level",       stats.get("threat_level","—"), "—"],
        ["Hybrid Decision",    primary.get("hybrid_decision") or "N/A", "—"],
        ["ML Confidence",      f"{ml_conf*100:.1f}%" if ml_conf else "N/A", "—"],
        ["Breach Status",      "Compromised" if primary.get("is_breached") else "Safe", "—"],
    ]
    for i, row in enumerate(kpi_rows, 6):
        ws1.append(row)
        _style_data_row(ws1, i, alt=(i % 2 == 0))

    _autofit(ws1)
    ws1.freeze_panes = "A6"

    # ── Sheet 2: Password Analysis ─────────────────────────────────────
    ws2 = wb.create_sheet("Password Analysis")
    ws2.sheet_view.showGridLines = False
    headers2 = ["#", "Label", "Category", "Score", "Entropy",
                 "Length", "Uppercase", "Lowercase", "Digits", "Special",
                 "Dictionary", "Keyboard", "Breached", "Analyzed At"]
    _write_headers(ws2, headers2)
    for i, r in enumerate(analyses, 2):
        row = [
            i-1, r.get("label","—"), r.get("strength_category","—"),
            r.get("strength_score","—"), r.get("entropy","—"),
            r.get("length","—"), r.get("uppercase",0), r.get("lowercase",0),
            r.get("digits",0), r.get("special",0),
            "Yes" if r.get("dictionary_word") else "No",
            "Yes" if r.get("keyboard_pattern") else "No",
            "Yes" if r.get("is_breached") else "No",
            r.get("analyzed_at","—"),
        ]
        ws2.append(row)
        _style_data_row(ws2, i, alt=(i%2==0), danger=bool(r.get("is_breached")))

    _autofit(ws2)
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Hybrid AI ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Hybrid AI")
    ws3.sheet_view.showGridLines = False
    headers3 = ["#", "Label", "Rule Engine", "ML Prediction",
                 "ML Confidence", "Hybrid Decision", "Agreement",
                 "Decision Source", "AI Reason"]
    _write_headers(ws3, headers3)
    hybrid_rows = [r for r in analyses if r.get("hybrid_decision")]
    for i, r in enumerate(hybrid_rows, 2):
        conf = r.get("ml_confidence")
        agr  = r.get("hybrid_agreement")
        ws3.append([
            i-1, r.get("label","—"),
            r.get("strength_category","—"),
            r.get("ml_prediction") or "—",
            f"{conf*100:.1f}%" if conf is not None else "—",
            r.get("hybrid_decision") or "—",
            "Yes" if agr is True else ("No" if agr is False else "—"),
            r.get("decision_source") or "—",
            r.get("hybrid_reason") or "—",
        ])
        _style_data_row(ws3, i, alt=(i%2==0))

    _autofit(ws3)
    ws3.freeze_panes = "A2"

    # ── Sheet 4: Breach Intelligence ───────────────────────────────────
    ws4 = wb.create_sheet("Breach Intelligence")
    ws4.sheet_view.showGridLines = False
    headers4 = ["#", "Label", "Strength", "Score", "Entropy",
                 "Breach Status", "Source", "Risk Level"]
    _write_headers(ws4, headers4)
    for i, r in enumerate(analyses, 2):
        breach = r.get("is_breached", False)
        ws4.append([
            i-1, r.get("label","—"),
            r.get("strength_category","—"),
            r.get("strength_score","—"),
            r.get("entropy","—"),
            "Breached" if breach else "Safe",
            "RockYou Sample" if breach else "—",
            "Critical" if breach else "Safe",
        ])
        _style_data_row(ws4, i, alt=(i%2==0), danger=breach)

    _autofit(ws4)
    ws4.freeze_panes = "A2"

    # ── Sheet 5: Analytics Summary ─────────────────────────────────────
    ws5 = wb.create_sheet("Analytics Summary")
    ws5.sheet_view.showGridLines = False
    ws5["A1"] = "Analytics Summary"
    ws5["A1"].font = _title_font()
    ws5.append([])
    ws5.append(["Metric", "Value"])
    for cell in ws5[3]:
        cell.font = _hdr_font(); cell.fill = _hdr_fill()
        cell.border = _thin_border()

    # Strength distribution
    cats   = Counter(r.get("strength_category","?") for r in analyses)
    total  = len(analyses)
    for cat_name in ["Very Weak","Weak","Moderate","Strong","Excellent"]:
        cnt = cats.get(cat_name, 0)
        pct = f"{cnt/total*100:.1f}%" if total else "0%"
        ws5.append([cat_name, f"{cnt}  ({pct})"])

    ws5.append(["—", "—"])
    ws5.append(["Total Analyses",  total])
    ws5.append(["Avg Score",       stats.get("avg_score", 0)])
    ws5.append(["Strong %",        f"{stats.get('strong_pct',0)}%"])
    ws5.append(["Weak %",          f"{stats.get('weak_pct',0)}%"])
    ws5.append(["Breach Rate %",   f"{stats.get('breached_pct',0)}%"])

    for i in range(4, ws5.max_row+1):
        _style_data_row(ws5, i, alt=(i%2==0))

    _autofit(ws5)

    # ── Sheet 6: Recommendations ───────────────────────────────────────
    ws6 = wb.create_sheet("Recommendations")
    ws6.sheet_view.showGridLines = False
    ws6["A1"] = "AI Security Recommendations"
    ws6["A1"].font = _title_font()
    ws6.append([])
    _write_headers(ws6, ["Category", "Recommendation"])

    recs = [
        ("Immediate",       "Change any breached passwords immediately on all services."),
        ("Immediate",       "Enable MFA on email and banking accounts today."),
        ("Password Hygiene","Use 14+ character passwords with all four character classes."),
        ("Password Hygiene","Never reuse passwords across different services."),
        ("Password Hygiene","Avoid dictionary words, keyboard patterns, and personal dates."),
        ("Credential Mgmt", "Store all passwords in an encrypted password manager."),
        ("Credential Mgmt", "Rotate high-value passwords every 90 days."),
        ("MFA",             "Use TOTP apps (Authy, Google Authenticator) over SMS OTP."),
        ("MFA",             "Deploy hardware security keys for enterprise administrator accounts."),
        ("Monitoring",      "Run weekly breach scans through SecureGuard AI."),
        ("Monitoring",      "Subscribe to HaveIBeenPwned notifications for new breach datasets."),
        ("Long-Term",       "Conduct quarterly credential audits across all organisational systems."),
        ("Long-Term",       "Train staff on phishing, social engineering, and credential hygiene."),
    ]
    for i, (cat, rec) in enumerate(recs, 4):
        ws6.append([cat, rec])
        _style_data_row(ws6, i, alt=(i%2==0))

    _autofit(ws6)

    out = io.BytesIO()
    wb.save(out)
    log.info("Excel report generated  user=%s  sheets=6  analyses=%d",
             user_name, len(analyses))
    return out.getvalue()
